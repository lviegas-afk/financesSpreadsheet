import os
import sys
import time
import gspread
from gspread import Worksheet
from oauth2client.service_account import ServiceAccountCredentials
from collections import deque
import xlrd
import re
from dataclasses import dataclass
import openpyxl

from scripts.src.wrapper.Database import Database
from src.ServiceMapper import ServiceMapper
from src.wrapper.Database import Database
import locale
from datetime import datetime, date
import os


@dataclass
class RowData:
    date: str
    desc: str
    referencia: str
    imp: str


class CustomWorkSheet:
    def __init__(self, sheet):
        self.sheet = sheet

    def getCellValue(self, row: int, col: int) -> str:
        if isinstance(self.sheet, openpyxl.worksheet.worksheet.Worksheet):
            value:str
            value = self.sheet.cell(row+1,col+1).value
            if value is None:
                return ''
            else:
                return value


def processXLSx(fileName: str, user: str):
    wb = openpyxl.load_workbook(os.path.join('excels', fileName))
    print(wb.sheetnames)
    ws: CustomWorkSheet
    ws = CustomWorkSheet(wb.active)
    print(type(ws))
    processSheet(ws)


def processXLS(fileName: str, user: str):
    wb = xlrd.open_workbook(os.path.join('excels', fileName))
    sheet = wb.sheet_by_index(0)
    processSheet(sheet)


def processSheet(sheet: CustomWorkSheet):
    col = 1
    row = 14
    print(sheet.getCellValue(row, col))
    cell = sheet.cell_value(row, col)
    de = deque()
    dbwrapper: Database = Database(os.getenv("APPDBPATH"))
    while cell != '':
        data = RowData(sheet.cell_value(row, col), sheet.cell_value(row, 3), sheet.cell_value(row, 4),
                       sheet.cell_value(row, 5))
        date_data = datetime.strptime(data.date, "%d/%m/%Y").date()
        importe = locale.atof(data.imp)
        result = dbwrapper.select_operation([data.referencia, importe, date_data, user])
        if not len(result):
            de.append(data)
        else:
            print('ya esta repetido')
        print('fecha: {} - descripcion: {} - importe: {}'.format(data['date'], data['desc'], data['imp']))
        row += 1
        cell = sheet.cell_value(row, col)

    cells = []
    grow = 2
    while de:
        data = de.pop()
        date_data = datetime.strptime(data.date, "%d/%m/%Y").date()
        importe = locale.atof(data.imp)
        serviceMapper = ServiceMapper(data.desc)
        jsoncell = [
            (date_data - date(1899, 12, 30)).days,
            data.desc,
            importe,
            serviceMapper.getDescription(),
            serviceMapper.getCategory(),
            user
        ]
        cells.append(jsoncell)
        dbwrapper.insert_operation([data.referencia, importe, date_data, serviceMapper.getCategory(), user])
    try:
        scope = ['https://www.googleapis.com/auth/spreadsheets', "https://www.googleapis.com/auth/drive"]
        credentials = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)
        client = gspread.authorize(credentials)
        worksheet = client.open('financesSpreadsheet').get_worksheet_by_id(0)
        worksheet.append_rows(cells)
    except gspread.exceptions.APIError:
        print("me timeouteo el api espero y vuelvo a probar")
        time.sleep(30)


if __name__ == '__main__':
    locale.setlocale(locale.LC_ALL, '')
    excelsDir = os.scandir('excels')
    for excel in excelsDir:
        if os.path.splitext(excel.name)[1] == '.xlsx':
            processXLSx(fileName=excel.name, user='')
        if os.path.splitext(excel.name)[1] == '.xls':
            processXLS(fileName=excel.name, user='')
